#!/usr/bin/env python
import argparse
import csv
import json
import re
from pathlib import Path


TEXT_ENCODINGS = ["utf-8-sig", "utf-8", "gb18030", "cp936", "latin-1"]


def readText(path, maxChars):
    lastError = None
    for encoding in TEXT_ENCODINGS:
        try:
            text = Path(path).read_text(encoding=encoding, errors="strict")
            return text[:maxChars], encoding, None
        except UnicodeError as error:
            lastError = str(error)
    text = Path(path).read_text(encoding="utf-8", errors="replace")
    return text[:maxChars], "utf-8-replace", lastError


def detectDelimiter(lines):
    candidates = [",", "\t", ";"]
    bestDelimiter = None
    bestScore = 0
    for delimiter in candidates:
        counts = [line.count(delimiter) for line in lines[:20] if line.strip()]
        if counts:
            score = max(counts)
            if score > bestScore:
                bestScore = score
                bestDelimiter = delimiter
    if bestDelimiter:
        return bestDelimiter
    return "whitespace"


def parseDelimited(text, delimiter, maxRows):
    lines = [line for line in text.splitlines() if line.strip()]
    if delimiter == "whitespace":
        rows = [re.split(r"\s+", line.strip()) for line in lines[:maxRows]]
    else:
        rows = list(csv.reader(lines[:maxRows], delimiter=delimiter))
    header = rows[0] if rows else []
    return {
        "delimiter": delimiter,
        "headerCandidate": header,
        "sampleRows": rows[1:maxRows] if len(rows) > 1 else [],
        "columnCount": max((len(row) for row in rows), default=0),
        "rowSampleCount": len(rows)
    }


def summarizeJson(path):
    text, encoding, warning = readText(path, 2_000_000)
    data = json.loads(text)
    if isinstance(data, dict):
        summary = {
            "jsonType": "object",
            "topLevelKeys": list(data.keys())[:100]
        }
    elif isinstance(data, list):
        firstItem = data[0] if data else None
        summary = {
            "jsonType": "array",
            "length": len(data),
            "firstItemType": type(firstItem).__name__
        }
    else:
        summary = {
            "jsonType": type(data).__name__
        }
    summary["encoding"] = encoding
    if warning:
        summary["warning"] = warning
    return summary


def summarizeXlsx(path, maxRows):
    try:
        import openpyxl
    except ImportError:
        return {
            "warning": "openpyxl is not installed; cannot read xlsx content.",
            "sheets": []
        }

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheetName in workbook.sheetnames[:10]:
        sheet = workbook[sheetName]
        rows = []
        for rowIndex, row in enumerate(sheet.iter_rows(values_only=True)):
            if rowIndex >= maxRows:
                break
            rows.append([cell for cell in row])
        sheets.append({
            "sheetName": sheetName,
            "sampleRows": rows,
            "maxRow": sheet.max_row,
            "maxColumn": sheet.max_column
        })
    return {"sheets": sheets}


def summarizeFile(filePath, maxRows, maxChars):
    path = Path(filePath)
    extension = path.suffix.lower()
    result = {
        "file": str(path.resolve()),
        "extension": extension,
        "size": path.stat().st_size,
        "warnings": []
    }

    if extension == ".json":
        result.update(summarizeJson(path))
        return result

    if extension in [".xlsx", ".xlsm"]:
        result.update(summarizeXlsx(path, maxRows))
        return result

    text, encoding, warning = readText(path, maxChars)
    if warning:
        result["warnings"].append(warning)
    lines = text.splitlines()
    delimiter = detectDelimiter(lines)
    result.update(parseDelimited(text, delimiter, maxRows))
    result["encoding"] = encoding
    result["textPreview"] = "\n".join(lines[:20])
    return result


def main():
    parser = argparse.ArgumentParser(description="Summarize a table-like motor project file as JSON.")
    parser.add_argument("filePath", help="File path to summarize.")
    parser.add_argument("--max-rows", type=int, default=30, help="Maximum sample rows.")
    parser.add_argument("--max-chars", type=int, default=200000, help="Maximum characters to read from text files.")
    parser.add_argument("--output", help="Optional JSON output path.")
    args = parser.parse_args()

    result = summarizeFile(args.filePath, args.max_rows, args.max_chars)
    outputText = json.dumps(result, ensure_ascii=False, indent=2, default=str)

    if args.output:
        Path(args.output).write_text(outputText, encoding="utf-8")
    else:
        print(outputText)


if __name__ == "__main__":
    main()
